import logging
import os

from aiogram.dispatcher import FSMContext
from openpyxl import Workbook

from aiogram import types

from handlers.incoming.function_inc import check_summary_main_inc
from keyboards.inline.out_in_keys import back_download
from loader import bot, db, dp
from states.user_states import PayHistoryIncoming


@dp.callback_query_handler(text="download-inc", state="*")
async def di_download(call: types.CallbackQuery):
    try:
        await call.message.delete()
        user_id = call.from_user.id
        category = await db.get_userall_inc(user_id=user_id)
        all_summary = await db.summary_all_inc(user_id=user_id)

        wb = Workbook()
        ws = wb.active

        ws.append(["Kirim", "Nomi", "Sana", "Summa (so'm)"])

        for data in category:
            ws.append(["Kirim", data[2], data[5], data[3]])
        ws.append(["Jami:", "", "", f"{all_summary}"])
        last_row = ws.max_row
        ws.merge_cells(f"A{last_row}:C{last_row}")
        wb.save("Kirim_IqtisodchiRobot.xlsx")

        await bot.send_document(chat_id=user_id,
                                document=types.InputFile(path_or_bytesio="Kirim_IqtisodchiRobot.xlsx"),
                                reply_markup=back_download)

        os.remove("Kirim_IqtisodchiRobot.xlsx")
        await PayHistoryIncoming.back_main.set()

    except Exception as err:
        logging.error(err)


@dp.callback_query_handler(state=PayHistoryIncoming.back_main)
async def di_back_incmain(call: types.CallbackQuery, state: FSMContext):

    await call.message.delete()

    await check_summary_main_inc(
        user_id=call.from_user.id,
        callback=call,
        no_edit=True,
        currency="so'm",
        section_name="📥 Kirim",
        total="📥 Kirim uchun jami:"
    )
    await state.finish()

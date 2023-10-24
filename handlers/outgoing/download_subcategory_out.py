import logging
import os

from openpyxl import Workbook

from aiogram import types
from aiogram.dispatcher import FSMContext

from handlers.all.menu_handlers import navigate
from states.user_states import DownloadHistoryOut

from keyboards.inline.out_keyboards import back_download, main_menu, categories_keyboard
from loader import bot, db, dp


@dp.callback_query_handler(text_contains="downloadcategory_", state="*")
async def dso_category_one(call: types.CallbackQuery, state: FSMContext):
    try:
        await call.message.delete()

        user_id = call.from_user.id
        category = call.data.split("_")[1]

        subcategory = await db.getdate_category_out(user_id=user_id, category_name=category)
        category_summary = await db.get_sum_category(user_id=user_id, category_name=category)

        if category_summary == 0:
            await call.answer(
                text="To'lovlar mavjud emas!",
                show_alert=True
            )

        else:
            wb = Workbook()

            ws = wb.active

            ws.append(["Bo'lim nomi", "Kategoriya nomi", "Summa (so'm)"])

            for data in subcategory:
                summary = await db.get_sum_subcategory(user_id=user_id, subcategory_name=data[0])

                ws.append(["Chiqim", data[0], summary])

            ws.append(["Jami:", " ", f"{category_summary}"])
            last_row = ws.max_row
            ws.merge_cells(f"A{last_row}:B{last_row}")

            wb.save("Chiqim_Kategoriya_@IqtisodchiRobot.xlsx")

            await bot.send_document(chat_id=user_id,
                                    document=types.InputFile(path_or_bytesio="Chiqim_Kategoriya_@IqtisodchiRobot.xlsx"),
                                    reply_markup=back_download)
            await call.message.answer(text="Ma'lumotlar Excel jadval shaklida yuborildi!")

            os.remove("Chiqim_Kategoriya_@IqtisodchiRobot.xlsx")

            await DownloadHistoryOut.subcategory.set()
    except Exception as err:
        logging.error(err)


@dp.callback_query_handler(state=DownloadHistoryOut.subcategory)
async def dso_back_menu(call: types.CallbackQuery, state: FSMContext):
    await call.message.delete()
    data = await state.get_data()

    await navigate(call=call,
                   callback_data=data,
                   state=state)
    await state.finish()

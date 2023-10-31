import logging
import os

from openpyxl import Workbook

from aiogram import types
from aiogram.dispatcher import FSMContext

from keyboards.inline.out_in_keys import back_download
from states.user_states import DownloadHistoryOut

from keyboards.inline.outgoing_keyboards import categories_keyboard
from loader import bot, db, dp


@dp.callback_query_handler(text="downloadall", state="*")
async def dco_category_one(call: types.CallbackQuery):
    try:
        await call.message.delete()

        user_id = call.from_user.id
        category = await db.get_categories_out(user_id=user_id, all_data=True)
        all_summary = await db.get_sum_all_out(user_id=user_id)

        if all_summary == 0:
            await call.answer(
                text="To'lovlar mavjud emas!",
                show_alert=True
            )
        else:
            wb = Workbook()

            ws = wb.active

            ws.append(["Chiqim", "Kategoriya", "Subkategoriya", "Sana", "Summa (so'm)"])

            for data in category:

                ws.append(["Chiqim", data[2], data[3], data[5], data[4]])

            ws.append(["Jami:", " ", " ", " ", f"{all_summary}"])
            last_row = ws.max_row
            ws.merge_cells(f"A{last_row}:D{last_row}")

            wb.save("Chiqim_IqtisodchiRobot.xlsx")

            await bot.send_document(chat_id=user_id,
                                    document=types.InputFile(path_or_bytesio="Chiqim_IqtisodchiRobot.xlsx"),
                                    reply_markup=back_download)

            os.remove("Chiqim_IqtisodchiRobot.xlsx")

            await DownloadHistoryOut.category.set()
    except Exception as err:
        logging.error(err)


@dp.callback_query_handler(state=DownloadHistoryOut.category)
async def dco_back_menu(call: types.CallbackQuery, state: FSMContext):

    await call.message.delete()
    await call.message.answer(
        text="<b>📤 Chiqim</b>",
        reply_markup=await categories_keyboard(user_id=call.from_user.id)
    )
    await state.finish()
